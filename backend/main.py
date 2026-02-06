from fastapi import FastAPI, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import select

from db.session import get_db, engine
from db.base import Base

from schemas.couple import (
    CoupleSessionCreate,
    CoupleSessionCreated,
    SubmitAnswers,
    SubmitMaleResponse,
    SubmitFinalResponse,
)
from schemas.question import QuestionsResponse, QuestionOut

from models.couple_result import CoupleResult
from models.couple_session import CoupleSession
from crud.couple import (
    create_couple_session,
    get_couple_session,
    get_or_assign_questions,
    submit_player_answers,
)

from fastapi.middleware.cors import CORSMiddleware

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from datetime import datetime


app = FastAPI(title="Couple Quiz Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,  # kalau allow_origins ["*"], ini harus False
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create tables
Base.metadata.create_all(bind=engine)


@app.post("/couple-sessions", response_model=CoupleSessionCreated)
def create_session(payload: CoupleSessionCreate, db: Session = Depends(get_db)):
    # print("male_name:", payload.male_name)
    # print("female_name:", payload.female_name)
    cs = create_couple_session(db, payload.male_name.strip(), payload.female_name.strip())
    # print(payload.model_dump())
    print(f"couple_session id : {cs.id}")
    return {"id": cs.id, "current_player": "male"}



@app.get("/couple-sessions/{couple_session_id}/questions", response_model=QuestionsResponse)
def get_questions(
    couple_session_id: int,
    player: str = Query(..., pattern="^(male|female)$"),
    db: Session = Depends(get_db),
):
    cs = get_couple_session(db, couple_session_id)
    if not cs:
        raise HTTPException(status_code=404, detail="Session not found")

    # enforce flow
    if player == "male":
        if cs.status not in ("created", "male_done"):
            raise HTTPException(status_code=400, detail="Invalid session status for male questions")
    if player == "female":
        if cs.status != "male_done":
            raise HTTPException(status_code=400, detail="Female can start only after male finished")

    questions = get_or_assign_questions(db, couple_session_id, player, n=5)
    return {
        "questions": [
            QuestionOut(
                id=q.id,
                question_text=q.question_text,
                option_a=q.option_a,
                option_b=q.option_b,
                correct_answer=q.correct_answer,
            )
            for q in questions
        ]
    }


@app.post("/couple-sessions/{couple_session_id}/submit")
def submit_answers(
    couple_session_id: int,
    payload: SubmitAnswers,
    db: Session = Depends(get_db),
):
    cs: CoupleSession | None = get_couple_session(db, couple_session_id)
    if not cs:
        raise HTTPException(status_code=404, detail="Session not found")

    player = payload.player.strip().lower()
    if player not in ("male", "female"):
        raise HTTPException(status_code=400, detail="player must be 'male' or 'female'")

    # enforce flow
    if player == "male":
        if cs.status != "created":
            raise HTTPException(status_code=400, detail="Male can submit only when status=created")
    if player == "female":
        if cs.status != "male_done":
            raise HTTPException(status_code=400, detail="Female can submit only after male finished")

    # submit & score
    try:
        score = submit_player_answers(
            db,
            couple_session_id,
            player,
            answers=[{"question_id": a.question_id, "answer": a.answer} for a in payload.answers],
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # SIMPAN SCORE KE couple_sessions
    if player == "male":
        cs.male_score = score
        cs.status = "male_done"
        db.add(cs)
        db.commit()
        db.refresh(cs)
        return SubmitMaleResponse(player="male", score=score, next_player="female")

    # female final
    cs.female_score = score
    cs.status = "done"
    cs.completed_at = datetime.utcnow()  # timestamp selesai
    db.add(cs)
    db.commit()
    db.refresh(cs)

    # UPSERT ke couple_results (1 session 1 result)
    existing = db.scalar(
        select(CoupleResult).where(CoupleResult.couple_session_id == cs.id)
    )
    if existing:
        existing.male_name = cs.male_name
        existing.female_name = cs.female_name
        existing.male_score = int(cs.male_score or 0)
        existing.female_score = int(cs.female_score or 0)
        existing.total_players = 2
        existing.created_at = cs.created_at
        existing.completed_at = cs.completed_at
    else:
        db.add(
            CoupleResult(
                couple_session_id=cs.id,
                male_name=cs.male_name,
                female_name=cs.female_name,
                male_score=int(cs.male_score or 0),
                female_score=int(cs.female_score or 0),
                total_players=2,
                created_at=cs.created_at,
                completed_at=cs.completed_at,
            )
        )

    db.commit()

    return SubmitFinalResponse(
        male_score=int(cs.male_score or 0),
        female_score=int(cs.female_score or 0),
        status="done",
    )

@app.get("/couple-sessions/{couple_session_id}")
def get_session_status(couple_session_id: int, db: Session = Depends(get_db)):
    cs = get_couple_session(db, couple_session_id)
    if not cs:
        raise HTTPException(status_code=404, detail="Session not found")

    return {
        "id": cs.id,
        "male_name": cs.male_name,
        "female_name": cs.female_name,
        "status": cs.status,
        "male_score": cs.male_score,
        "female_score": cs.female_score,
        "created_at": cs.created_at,
        "completed_at": cs.completed_at,
    }


@app.get("/reports/couple-results.xlsx")
def export_couple_results_xlsx(db: Session = Depends(get_db)):
    # Ambil semua result, urut terbaru
    rows = db.execute(
        select(CoupleResult).order_by(CoupleResult.completed_at.desc())
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header kolom
    headers = [
        "result_id",
        "couple_session_id",
        "male_name",
        "female_name",
        "male_score",
        "female_score",
        "total_players",
        "created_at",
        "completed_at",
    ]
    ws.append(headers)

    # Isi data
    for r in rows:
        ws.append([
            r.id,
            r.couple_session_id,
            r.male_name,
            r.female_name,
            r.male_score,
            r.female_score,
            r.total_players,
            r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
            r.completed_at.strftime("%Y-%m-%d %H:%M:%S") if r.completed_at else "",
        ])

    # (Optional) auto width sederhana biar enak dibaca
    for col_idx, col_name in enumerate(headers, start=1):
        max_len = len(col_name)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Simpan ke memory
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"couple_results_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
